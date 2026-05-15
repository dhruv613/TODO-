"""Excel-backed user store + password hashing."""
import hashlib
import secrets
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook

BASE_DIR = Path(__file__).resolve().parent
USERS_FILE = BASE_DIR / "users.xlsx"
HEADERS = ["id", "username", "password_hash", "created_at"]


def init_users_file() -> None:
    if USERS_FILE.exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Users"
    ws.append(HEADERS)
    # Width tweaks so the file looks tidy when opened in Excel
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 90
    ws.column_dimensions["D"].width = 26
    wb.save(USERS_FILE)


def _hash(password: str, salt: str) -> str:
    digest = hashlib.pbkdf2_hmac("sha256", password.encode(), salt.encode(), 120_000)
    return f"{salt}${digest.hex()}"


def hash_password(password: str) -> str:
    return _hash(password, secrets.token_hex(16))


def verify_password(password: str, stored: str) -> bool:
    try:
        salt, _ = stored.split("$", 1)
    except ValueError:
        return False
    return secrets.compare_digest(_hash(password, salt), stored)


def _load_rows():
    wb = load_workbook(USERS_FILE)
    return wb, wb.active


def list_users() -> list[dict]:
    _, ws = _load_rows()
    users = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        users.append({"id": row[0], "username": row[1], "password_hash": row[2], "created_at": row[3]})
    return users


def find_user(username: str) -> dict | None:
    username = (username or "").strip().lower()
    for user in list_users():
        if user["username"].lower() == username:
            return user
    return None


def create_user(username: str, password: str) -> dict | None:
    username = username.strip()
    if find_user(username):
        return None
    user_id = "u_" + secrets.token_hex(6)
    pwd_hash = hash_password(password)
    created = datetime.now().isoformat(timespec="seconds")
    wb, ws = _load_rows()
    ws.append([user_id, username, pwd_hash, created])
    wb.save(USERS_FILE)
    return {"id": user_id, "username": username, "created_at": created}


def authenticate(username: str, password: str) -> dict | None:
    user = find_user(username)
    if not user:
        return None
    if not verify_password(password, user["password_hash"]):
        return None
    return {"id": user["id"], "username": user["username"]}
